import sys
import os
from openpyxl import load_workbook
from shutil import copyfile
import pandas as pd
from pathlib import Path
import pickle
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

os.chdir('python_backend')
SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']

TOKEN_FILE = 'token.pickle'

CREDENTIALS_FILE = 'credentials.json'

def start(path, current_client):
    pdf_path = path
    file_nameid = (Path(pdf_path)).stem
    if current_client == 'Caterpillar':
        convert(file_nameid)
        upload(file_nameid)
    else:
        generic_convert(file_nameid)

def convert(file_nameid):
    caminho_pasta_excel = 'Excel'
    excel_final = 'planilha_final.xlsx'
    excel_modelo = 'cat_importar_dados.xlsx'
    copyfile(excel_modelo, excel_final)
    arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
    wb_destino = load_workbook(excel_final)
    c = 1
    p = 1
    w = 1
    for arquivo in arquivos_excel:
        if file_nameid in arquivo:
            wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
            table_name = 'tabela'

            if arquivo != 'planilha_final.xlsx':
                string1 = 'CIRCUIT DATA TABLE'
                string2 = 'PARTS LIST'
                string3 = 'BUNDLE TABLE'

                df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                file_path_csv = 'arquivo.csv'
                df_paracsv.to_csv(file_path_csv, index=False)
                df_readcsv = pd.read_csv(file_path_csv, header=None, skiprows=1)
                with open(file_path_csv, 'r', encoding='utf-8') as file:
                    file_content = file.read()
                
                if(df_readcsv == string1).any().any():
                    table_name = f'Circuit{c}'
                elif string1 in file_content:
                    table_name = f'Circuit{c}'
                
                if ((df_paracsv == string2)).any().any():
                    table_name = f'Partlist{p}'
                    p += 1
                elif string2 in file_content:
                    table_name = f'Partlist{p}'
                    p += 1    

                if(df_readcsv == string3).any().any():
                    table_name = f'Bundle{w}'
                    w += 1
                elif string3 in file_content:
                        table_name = f'Bundle{w}'
                        w += 1

                for sheet_name in wb_origem.sheetnames:
                    ws_origem = wb_origem[sheet_name]
                    ws_destino = wb_destino.create_sheet(title=table_name)
                    for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                        ws_destino.append(row)

                    for row in ws_origem.iter_rows(min_row=2, values_only=True):
                        ws_destino.append(row)

        arquivo_excel_path = f'Excel/planilha_final{file_nameid}.xlsx'
        wb_destino.save(arquivo_excel_path)
        df_final = pd.read_excel(arquivo_excel_path, sheet_name='DADOS')
        df_final.at[3, 'X'] = c
        df_final.at[18, 'X'] = p

def generic_convert(file_nameid):
    caminho_pasta_excel = 'Excel'
    excel_final = 'generic_spreadsheet.xlsx'
    arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
    wb_destino = load_workbook(excel_final)

    for arquivo in arquivos_excel:
        if file_nameid in arquivo:
            wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
            table_name = 'tabela'

            if arquivo != 'planilha_final.xlsx':

                df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                file_path_csv = 'arquivo.csv'
                df_paracsv.to_csv(file_path_csv, index=False)

                for sheet_name in wb_origem.sheetnames:
                    ws_origem = wb_origem[sheet_name]
                    ws_destino = wb_destino.create_sheet(title=table_name)
                    for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                        ws_destino.append(row)

                    for row in ws_origem.iter_rows(min_row=2, values_only=True):
                        ws_destino.append(row)

        arquivo_excel_path = f'Excel/planilha_final{file_nameid}.xlsx'
        wb_destino.save(arquivo_excel_path)

def upload(file_nameid):
    def authenticate():
        creds = None
        if os.path.exists(TOKEN_FILE):
            with open(TOKEN_FILE, 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, 'wb') as token:
                pickle.dump(creds, token)
        return creds

    def replace_spreadsheet(creds, local_file_path, destination_file_id):
        drive_service = build('drive', 'v3', credentials=creds)
        media = MediaFileUpload(local_file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().update(
            fileId=destination_file_id,
            media_body=media).execute()

    def main():
        creds = authenticate()
        local_file_path = f'Excel/planilha_final{file_nameid}.xlsx'
        destination_file_id = '1wnrxtfUl4DLw9uEpudIdR9NCSR8XDYY0bW4tAWsmbEk'
        replace_spreadsheet(creds, local_file_path, destination_file_id)

    main()

if __name__ == '__main__':
    arg1 = sys.argv[1]
    arg2 = sys.argv[2]
    start(arg1, arg2)